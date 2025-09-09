# run_from_settings.py
from __future__ import annotations
import json
import os
import sys
import math
import subprocess
from pathlib import Path
from typing import Any, Dict, List

import pandas as pd

from spanner import (
    greedy_spanner,
    mst_weight,
    graph_weight,
    compute_stretch,
    random_complete_weighted_graph,
    er_random_graph,
    high_girth_random_graph,   # זמין בקובץ spanner.py המעודכן
)

def _open_with_default_app(path: str) -> None:
    """פתח קובץ באפליקציית ברירת המחדל של מערכת ההפעלה."""
    try:
        if sys.platform.startswith("win"):
            os.startfile(path)  # type: ignore[attr-defined]
        elif sys.platform == "darwin":
            subprocess.run(["open", path], check=False)
        else:
            subprocess.run(["xdg-open", path], check=False)
    except Exception as e:
        print(f"[warn] לא הצלחתי לפתוח אוטומטית את הקובץ: {e}")

def _validate_and_normalize_settings(cfg: Dict[str, Any]) -> Dict[str, Any]:
    # graph_type
    gtype = str(cfg.get("graph_type", "er")).lower()
    if gtype not in {"complete", "er", "high_girth"}:
        raise ValueError("graph_type חייב להיות אחד: 'complete' | 'er' | 'high_girth'")

    # n
    if "n" not in cfg:
        raise ValueError("נדרש שדה n (מספר הקודקודים)")
    n = int(cfg["n"])
    if n <= 0:
        raise ValueError("n חייב להיות מספר חיובי")

    # p (רק ל-ER)
    p = float(cfg.get("p", 0.2))
    if gtype == "er" and not (0.0 <= p <= 1.0):
        raise ValueError("ל-ER, הפרמטר p חייב להיות בין 0 ל-1")

    # t
    if "t" not in cfg:
        raise ValueError("נדרש שדה t (stretch factor)")
    t = float(cfg["t"])
    if t < 1.0:
        raise ValueError("t חייב להיות >= 1.0")

    # weight_range
    wr = cfg.get("weight_range", [0.0, 1.0])
    if not (isinstance(wr, list) and len(wr) == 2):
        raise ValueError("weight_range חייב להיות רשימה באורך 2: [low, high]")
    w_low, w_high = float(min(wr)), float(max(wr))
    if w_low == w_high:
        print("[warn] weight_range עם תחום ריק; כל המשקולות יהיו זהות")

    # runs_per_setting
    runs = int(cfg.get("runs_per_setting", 1))
    if runs <= 0:
        raise ValueError("runs_per_setting חייב להיות >= 1")

    # output
    out_cfg = cfg.get("output", {})
    csv_path = out_cfg.get("csv_path", "results/spanner_results.csv")
    excel_path = out_cfg.get("excel_path", "results/spanner_results.xlsx")
    open_after = bool(out_cfg.get("open_after", True))

    return {
        "graph_type": gtype,
        "n": n,
        "p": p,
        "t": t,
        "w_low": w_low,
        "w_high": w_high,
        "runs_per_setting": runs,
        "csv_path": csv_path,
        "excel_path": excel_path,
        "open_after": open_after,
    }

def _build_graph(graph_type: str, n: int, p: float, t: float, w_low: float, w_high: float):
    if graph_type == "complete":
        G = random_complete_weighted_graph(n, w_low=w_low, w_high=w_high)
        name = f"complete(n={n})"
    elif graph_type == "er":
        G = er_random_graph(n, p=p, w_low=w_low, w_high=w_high)
        name = f"er(n={n},p={p})"
    elif graph_type == "high_girth":
        # משתמש בפונקציה שלך שמייצרת p ~ n^(1/t) / n
        G = high_girth_random_graph(n, t=t, w_low=w_low, w_high=w_high)
        name = f"high_girth(n={n},t={t})"
    else:
        raise ValueError(f"סוג גרף לא נתמך: {graph_type}")
    return G, name

def run_from_settings(settings_path: str) -> None:
    # --- קריאה ואימות קובץ ההגדרות ---
    with open(settings_path, "r", encoding="utf-8") as f:
        raw = json.load(f)
    cfg = _validate_and_normalize_settings(raw)

    graph_type = cfg["graph_type"]
    n = cfg["n"]
    p = cfg["p"]
    t = cfg["t"]
    w_low = cfg["w_low"]
    w_high = cfg["w_high"]
    runs_per_setting = cfg["runs_per_setting"]
    csv_path = cfg["csv_path"]
    excel_path = cfg["excel_path"]
    open_after = cfg["open_after"]

    # תיקיות פלט
    Path(csv_path).parent.mkdir(parents=True, exist_ok=True)
    Path(excel_path).parent.mkdir(parents=True, exist_ok=True)

    # --- הרצות ואיסוף נתונים ---
    import time
    import random
    import numpy as np

    rows: List[Dict[str, Any]] = []
    for run_idx in range(1, runs_per_setting + 1):
        # זרע מגוון אך דטרמיניסטי יחסית להרצות שונות
        seed = (int(time.time() * 1e6) ^ (n << 16) ^ run_idx) & 0xFFFFFFFF
        random.seed(seed)
        np.random.seed(seed)

        G, gname = _build_graph(graph_type, n, p, t, w_low, w_high)
        H = greedy_spanner(G, t=t)

        n_nodes = G.number_of_nodes()
        m_edges = G.number_of_edges()
        w_mst = mst_weight(G)
        wH = graph_weight(H)
        stretch = compute_stretch(G, H)  # שים לב: all-pairs, יקר מאוד לגרפים גדולים

        rows.append({
            "run": run_idx,
            "graph_type": graph_type,
            "graph": gname,
            "n": n_nodes,
            "m": m_edges,
            "p": p if graph_type == "er" else None,
            "t": t,
            "weight_low": w_low,
            "weight_high": w_high,
            "spanner_edges": H.number_of_edges(),
            "spanner_weight": wH,
            "mst_weight": w_mst,
            "weight_ratio": (wH / w_mst) if w_mst > 0 else math.nan,
            "max_stretch": stretch,
            "seed": int(seed),
        })

    if not rows:
        print("לא נאספו נתונים. בדוק את קובץ ההגדרות.")
        return

    df = pd.DataFrame(rows)
    df.to_csv(csv_path, index=False, encoding="utf-8")
    print(f"[ok] CSV -> {csv_path}")

    # טבלת סיכום לפי (graph_type, n, t[, p]) עם ממוצעים/ת״ת
    group_cols = ["graph_type", "n", "t"]
    if graph_type == "er":
        group_cols.append("p")

    summary = (
        df.groupby(group_cols)
          .agg(
              runs=("run", "count"),
              edges_mean=("spanner_edges", "mean"),
              edges_std=("spanner_edges", "std"),
              wratio_mean=("weight_ratio", "mean"),
              wratio_std=("weight_ratio", "std"),
              stretch_mean=("max_stretch", "mean"),
              stretch_max=("max_stretch", "max"),
              mst_mean=("mst_weight", "mean"),
              spanner_w_mean=("spanner_weight", "mean"),
          )
          .reset_index()
    )

    # כתיבה ל-Excel עם שני גיליונות: raw + summary
    with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="raw_results")
        summary.to_excel(writer, index=False, sheet_name="summary")

        # התאמת רוחבי עמודות בסיסית
        from openpyxl.utils import get_column_letter
        wb = writer.book
        for sheet_name in ["raw_results", "summary"]:
            ws = wb[sheet_name]
            for col_idx, col in enumerate(ws.iter_cols(min_row=1, max_row=1), start=1):
                header = col[0].value or ""
                max_len = len(str(header))
                sample_rows = min(ws.max_row, 150)
                for cell in ws.iter_rows(min_row=2, max_row=sample_rows, min_col=col_idx, max_col=col_idx):
                    val = cell[0].value
                    if val is None:
                        continue
                    max_len = max(max_len, len(str(val)))
                ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 40)

    print(f"[ok] Excel -> {excel_path}")

    if open_after:
        target = excel_path if Path(excel_path).exists() else csv_path
        _open_with_default_app(target)

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("שימוש: python run_from_settings.py <path/to/settings.json>")
        sys.exit(1)
    run_from_settings(sys.argv[1])
